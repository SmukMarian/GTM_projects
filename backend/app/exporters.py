"""Экспорт данных в Excel."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import re
from typing import Iterable
from uuid import UUID, uuid4

from openpyxl import Workbook, load_workbook

from .models import (
    CharacteristicField,
    CharacteristicSection,
    ChecklistItem,
    Comment,
    FieldType,
    GTMStage,
    Project,
    ProjectStatus,
    PriorityLevel,
    ProductGroup,
    StageStatus,
    Subtask,
    Task,
    TaskStatus,
    TaskUrgency,
)


def _find_current_stage_name(project: Project) -> str | None:
    if not project.current_gtm_stage_id:
        return None
    for stage in project.gtm_stages:
        if stage.id == project.current_gtm_stage_id:
            return stage.title
    return None


CUSTOM_FIELD_PREFIX = "CF:"
EXCEL_SHEET_LIMIT = 31


def _make_sheet_name(base: str, used: set[str]) -> str:
    """Сформировать валидное имя листа и избежать дубликатов."""

    cleaned = re.sub(r"[\\/*?:\[\]]", "_", base).strip() or "Проект"
    trimmed = cleaned[:EXCEL_SHEET_LIMIT]
    candidate = trimmed
    suffix = 1
    while candidate in used:
        suffix_text = f"_{suffix}"
        candidate = (trimmed[: EXCEL_SHEET_LIMIT - len(suffix_text)] + suffix_text) or f"{cleaned[:EXCEL_SHEET_LIMIT-2]}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def export_projects_to_excel(
    *,
    projects: Iterable[Project],
    groups: Iterable[ProductGroup],
    statuses: set[ProjectStatus] | None = None,
    include_archived: bool = True,
    brand: str | None = None,
    current_stage_id: UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
) -> bytes:
    """Сформировать Excel-файл со списком проектов с полным набором полей."""

    projects_list: list[Project] = list(projects)
    if statuses is not None:
        projects_list = [p for p in projects_list if p.status in statuses]
    if not include_archived:
        projects_list = [p for p in projects_list if p.status != ProjectStatus.ARCHIVED]
    if brand:
        projects_list = [p for p in projects_list if p.brand.lower() == brand.lower()]
    if current_stage_id:
        projects_list = [p for p in projects_list if p.current_gtm_stage_id == current_stage_id]
    if planned_from:
        projects_list = [
            p
            for p in projects_list
            if p.planned_launch is not None and p.planned_launch >= planned_from
        ]
    if planned_to:
        projects_list = [
            p
            for p in projects_list
            if p.planned_launch is not None and p.planned_launch <= planned_to
        ]

    group_name_by_id = {group.id: group.name for group in groups}
    custom_keys: list[str] = sorted({key for p in projects_list for key in p.custom_fields.keys()})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Проекты"

    headers = [
        "ID",
        "Короткий ID",
        "Название проекта",
        "Продуктовая группа",
        "Бренд",
        "Рынок/регион",
        "Статус",
        "Плановая дата запуска",
        "Фактическая дата запуска",
        "Текущий GTM-этап",
        "Приоритет",
        "MOQ",
        "FOB",
        "PROMO",
        "RRP",
        "Краткое описание",
        "Полное описание",
    ]
    headers.extend(f"{CUSTOM_FIELD_PREFIX}{key}" for key in custom_keys)
    sheet.append(headers)

    for project in projects_list:
        current_stage_name = _find_current_stage_name(project)
        row = [
            str(project.id),
            project.short_id,
            project.name,
            group_name_by_id.get(project.group_id, ""),
            project.brand,
            project.market,
            project.status.value,
            project.planned_launch,
            project.actual_launch,
            current_stage_name,
            project.priority.value if project.priority else None,
            project.moq,
            project.fob_price,
            project.promo_price,
            project.rrp_price,
            project.short_description,
            project.full_description,
        ]
        for key in custom_keys:
            row.append(project.custom_fields.get(key))
        sheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_gtm_sheet(workbook: Workbook, project: Project, title: str = "GTM") -> None:
    sheet = workbook.create_sheet(title)

    headers = [
        "Порядок этапа",
        "Название этапа",
        "Описание этапа",
        "Плановая дата начала",
        "Плановая дата окончания",
        "Фактическая дата завершения",
        "Статус этапа",
        "Риск по этапу",
        "Чек-лист",
        "Порядок задачи",
        "Название задачи",
        "Описание задачи",
        "Статус задачи",
        "Срок задачи",
        "Важная задача",
        "Срочность задачи",
        "Порядок подзадачи",
        "Название подзадачи",
        "Подзадача выполнена",
        "Комментарий задачи",
        "Дата комментария",
    ]
    sheet.append(headers)

    ordered_stages = sorted(project.gtm_stages, key=lambda s: s.order)
    stage_order_map = {stage.id: idx for idx, stage in enumerate(ordered_stages, start=1)}
    ordered_tasks = sorted(
        project.tasks,
        key=lambda t: (
            stage_order_map.get(t.gtm_stage_id, 9999),
            getattr(t, "order", 0),
            (t.due_date or date.max),
            t.title.lower(),
        ),
    )

    for stage in ordered_stages:
        related_tasks = [t for t in ordered_tasks if t.gtm_stage_id == stage.id] or [None]
        stage_checklist = "; ".join(
            f"{'[x]' if item.done else '[ ]'} {item.title}" for item in sorted(stage.checklist, key=lambda c: c.order)
        )

        for task_idx, task in enumerate(related_tasks, start=1):
            subtasks = sorted(task.subtasks, key=lambda s: s.order) if task else []
            comments = task.comments if task else []
            rows = max(1, len(subtasks), len(comments))
            for row_idx in range(rows):
                sub = subtasks[row_idx] if row_idx < len(subtasks) else None
                comment = comments[row_idx] if row_idx < len(comments) else None
                sheet.append(
                    [
                        stage.order,
                        stage.title,
                        stage.description,
                        stage.planned_start,
                        stage.planned_end,
                        stage.actual_end,
                        stage.status.value,
                        "да" if stage.risk_flag else "нет",
                        stage_checklist,
                        task_idx if task else None,
                        task.title if task else None,
                        task.description if task else None,
                        task.status.value if task else None,
                        task.due_date if task else None,
                        "да" if (task and task.important) else None,
                        task.urgency.value if task else None,
                        sub.order if sub else None,
                        sub.title if sub else None,
                        "да" if (sub and sub.done) else None,
                        comment.text if comment else None,
                        (comment.created_at.isoformat() if comment else None),
                    ]
                )


def export_gtm_stages_to_excel(project: Project) -> bytes:
    """Выгрузить этапы, задачи, подзадачи и комментарии в одну таблицу."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_gtm_sheet(workbook, project, "GTM")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_characteristics_sheet(workbook: Workbook, project: Project, title: str = "Характеристики") -> None:
    sheet = workbook.create_sheet(title)

    headers = [
        "Секция",
        "Порядок секции",
        "Label RU",
        "Label EN",
        "Value RU",
        "Value EN",
        "Тип поля",
        "Порядок поля",
    ]
    sheet.append(headers)

    for section in sorted(project.characteristics, key=lambda s: s.order):
        for field in sorted(section.fields, key=lambda f: f.order):
            sheet.append(
                [
                    section.title,
                    section.order,
                    field.label_ru,
                    field.label_en,
                    field.value_ru,
                    field.value_en,
                    field.field_type.value,
                    field.order,
                ]
            )


def export_characteristics_to_excel(project: Project) -> bytes:
    """Сформировать Excel-файл с характеристиками проекта."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_characteristics_sheet(workbook, project, "Характеристики")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_project_sheet(
    workbook: Workbook, project: Project, title: str = "Проект", group_lookup: dict[UUID, str] | None = None
) -> None:
    sheet = workbook.create_sheet(title)

    custom_keys = sorted(project.custom_fields.keys())
    headers = [
        "ID",
        "Короткий ID",
        "Название проекта",
        "Продуктовая группа",
        "Бренд",
        "Рынок/регион",
        "Статус",
        "Плановая дата запуска",
        "Фактическая дата запуска",
        "Текущий GTM-этап",
        "Приоритет",
        "MOQ",
        "FOB",
        "PROMO",
        "RRP",
        "Краткое описание",
        "Полное описание",
    ]
    headers.extend(f"{CUSTOM_FIELD_PREFIX}{key}" for key in custom_keys)
    sheet.append(headers)

    current_stage_name = _find_current_stage_name(project)
    row = [
        str(project.id),
        project.short_id,
        project.name,
        group_lookup.get(project.group_id, project.group_id) if group_lookup else project.group_id,
        project.brand,
        project.market,
        project.status.value,
        project.planned_launch,
        project.actual_launch,
        current_stage_name,
        project.priority.value if project.priority else None,
        project.moq,
        project.fob_price,
        project.promo_price,
        project.rrp_price,
        project.short_description,
        project.full_description,
    ]
    for key in custom_keys:
        row.append(project.custom_fields.get(key))
    sheet.append(row)


def export_project_bundle(project: Project, groups: Iterable[ProductGroup] | None = None) -> bytes:
    """Выгрузить проект с базовыми полями, характеристиками и GTM в один файл."""

    workbook = Workbook()
    workbook.remove(workbook.active)

    group_lookup = {g.id: g.name for g in groups} if groups else {}

    _write_project_sheet(workbook, project, "Основные параметры", group_lookup)
    _write_characteristics_sheet(workbook, project, "Характеристики")
    _write_gtm_sheet(workbook, project, "GTM")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _parse_bool(value: str | bool | None) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    normalized = str(value).strip().lower()
    return normalized in {"1", "true", "да", "y", "yes", "oui", "on"}


STATUS_ALIASES = {
    "не начат": StageStatus.NOT_STARTED,
    "not started": StageStatus.NOT_STARTED,
    "в работе": StageStatus.IN_PROGRESS,
    "in progress": StageStatus.IN_PROGRESS,
    "done": StageStatus.DONE,
    "завершен": StageStatus.DONE,
    "завершён": StageStatus.DONE,
    "отменен": StageStatus.CANCELLED,
    "отменён": StageStatus.CANCELLED,
    "cancelled": StageStatus.CANCELLED,
}

TASK_STATUS_ALIASES = {
    "todo": TaskStatus.TODO,
    "to do": TaskStatus.TODO,
    "в работе": TaskStatus.IN_PROGRESS,
    "in progress": TaskStatus.IN_PROGRESS,
    "делается": TaskStatus.IN_PROGRESS,
    "done": TaskStatus.DONE,
    "сделано": TaskStatus.DONE,
    "готово": TaskStatus.DONE,
}

URGENCY_ALIASES = {
    "normal": TaskUrgency.NORMAL,
    "обычная": TaskUrgency.NORMAL,
    "нормальная": TaskUrgency.NORMAL,
    "high": TaskUrgency.HIGH,
    "срочно": TaskUrgency.HIGH,
    "высокая": TaskUrgency.HIGH,
}


def import_gtm_stages_from_excel(content: bytes) -> tuple[list[GTMStage], list[Task], list[str]]:
    """Распарсить Excel с этапами GTM, задачами, подзадачами и комментариями."""

    try:
        workbook = load_workbook(filename=BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        return [], [], [f"Не удалось прочитать Excel: {exc}"]

    sheet = workbook.active
    return import_gtm_stages_from_sheet(sheet)


def import_gtm_stages_from_sheet(sheet) -> tuple[list[GTMStage], list[Task], list[str]]:
    try:
        first_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        return [], [], ["Файл Excel пуст"]
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in first_row]
    header_map = {title.lower(): idx for idx, title in enumerate(header_row) if title}

    def normalize(value: str | None) -> str:
        return value.strip().lower() if value else ""

    def col(key: str, default: int | None = None) -> int | None:
        if key in header_map:
            return header_map[key]
        return default

    required_columns = {"название этапа"}
    missing = required_columns - set(header_map)
    if missing:
        return [], [], [f"Отсутствуют обязательные столбцы: {', '.join(sorted(missing))}"]

    combined_headers = {"название задачи", "название подзадачи", "комментарий задачи"}
    if combined_headers & set(header_map):
        return _import_gtm_single_sheet(sheet, header_map)

    return _import_gtm_legacy(sheet.parent, header_map)


def _import_gtm_single_sheet(sheet, header_map: dict[str, int]):
    stages: list[GTMStage] = []
    tasks: list[Task] = []
    errors: list[str] = []

    stage_index: dict[str, GTMStage] = {}
    task_index: dict[tuple[UUID, str, int], Task] = {}

    def normalize(value: str | None) -> str:
        return value.strip().lower() if value else ""

    def col(key: str, default: int | None = None) -> int | None:
        if key in header_map:
            return header_map[key]
        return default

    def parse_datetime(value) -> datetime | None:
        if isinstance(value, datetime):
            return value
        if value is None:
            return None
        try:
            return datetime.fromisoformat(str(value))
        except ValueError:
            return None

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        stage_title_raw = row[col("название этапа")]
        if not stage_title_raw:
            errors.append(f"Строка {row_index}: пустое название этапа")
            continue

        stage_key = normalize(str(stage_title_raw))
        stage = stage_index.get(stage_key)
        if stage is None:
            order_raw = row[col("порядок этапа")]
            try:
                order_value = int(order_raw) if order_raw not in (None, "") else len(stages)
            except (TypeError, ValueError):
                order_value = len(stages)

            status_cell = col("статус этапа")
            status_raw = row[status_cell] if status_cell is not None else None
            status_value = StageStatus.NOT_STARTED
            if status_raw:
                normalized = str(status_raw).strip().lower()
                status_value = STATUS_ALIASES.get(normalized, None) or StageStatus.__members__.get(normalized.upper(), None)
                if status_value is None:
                    errors.append(f"Строка {row_index}: неизвестный статус '{status_raw}'")
                    status_value = StageStatus.NOT_STARTED

            checklist_cell = col("чек-лист")
            checklist_raw = row[checklist_cell] if checklist_cell is not None else None
            checklist_items: list[str] = []
            if checklist_raw:
                checklist_items = [part.strip() for part in str(checklist_raw).split(";") if part.strip()]

            checklist_models = []
            for idx, item in enumerate(checklist_items):
                text = item
                done = False
                if item.startswith("[x]"):
                    done = True
                    text = item[3:].strip()
                elif item.startswith("[ ]"):
                    text = item[3:].strip()
                checklist_models.append(ChecklistItem(title=text, done=done, order=idx))

            stage = GTMStage(
                title=str(stage_title_raw).strip(),
                description=row[col("описание этапа")],
                order=order_value,
                planned_start=row[col("плановая дата начала")],
                planned_end=row[col("плановая дата окончания")],
                actual_end=row[col("фактическая дата завершения")],
                status=status_value,
                risk_flag=_parse_bool(row[col("риск по этапу")]) if col("риск по этапу") is not None else False,
                checklist=checklist_models,
            )
            stages.append(stage)
            stage_index[stage_key] = stage

        task_title_raw = row[col("название задачи")]
        task_order_raw = row[col("порядок задачи")]
        task_key = None
        if task_title_raw:
            try:
                task_order = int(task_order_raw) if task_order_raw not in (None, "") else 0
            except (TypeError, ValueError):
                task_order = 0
            task_key = (stage.id, normalize(str(task_title_raw)), task_order)

            if task_key not in task_index:
                status_raw = row[col("статус задачи")]
                status = TaskStatus.TODO
                if status_raw:
                    status = TASK_STATUS_ALIASES.get(str(status_raw).strip().lower(), None) or TaskStatus.__members__.get(
                        str(status_raw).strip().upper(), TaskStatus.TODO
                    )

                urgency_raw = row[col("срочность задачи")]
                urgency = TaskUrgency.NORMAL
                if urgency_raw:
                    urgency = URGENCY_ALIASES.get(str(urgency_raw).strip().lower(), TaskUrgency.NORMAL)

                important_raw = row[col("важная задача")]
                important = _parse_bool(important_raw) if important_raw is not None else False

                task_obj = Task(
                    title=str(task_title_raw).strip(),
                    description=row[col("описание задачи")],
                    status=status,
                    due_date=row[col("срок задачи")],
                    important=important,
                    urgency=urgency,
                    gtm_stage_id=stage.id,
                    subtasks=[],
                    comments=[],
                )
                task_index[task_key] = task_obj
                tasks.append(task_obj)

            task_obj = task_index[task_key]

            sub_title = row[col("название подзадачи")]
            if sub_title:
                try:
                    sub_order = int(row[col("порядок подзадачи")]) if col("порядок подзадачи") is not None else 0
                except (TypeError, ValueError):
                    sub_order = 0
                done_raw = row[col("подзадача выполнена")]
                done = _parse_bool(done_raw)
                task_obj.subtasks.append(Subtask(title=str(sub_title).strip(), done=done, order=sub_order))

            comment_text = row[col("комментарий задачи")]
            if comment_text:
                created_at_col = col("дата комментария")
                created_at = parse_datetime(row[created_at_col]) if created_at_col is not None else None
                task_obj.comments.append(
                    Comment(text=str(comment_text).strip(), created_at=created_at or datetime.utcnow())
                )

    stages.sort(key=lambda s: s.order)
    stage_order_map = {stage.id: stage.order for stage in stages}
    for stage in stages:
        stage.checklist.sort(key=lambda c: c.order)
    tasks.sort(key=lambda t: (stage_order_map.get(t.gtm_stage_id, 999), t.title.lower()))
    for task in tasks:
        task.subtasks.sort(key=lambda s: s.order)
        
    return stages, tasks, errors


def _import_gtm_legacy(workbook, header_map: dict[str, int]):
    sheet = workbook.active
    stages: list[GTMStage] = []
    tasks_raw: list[tuple[int, GTMStage, Task]] = []
    errors: list[str] = []

    def normalize(value: str | None) -> str:
        return value.strip().lower() if value else ""

    def col(key: str, default: int | None = None) -> int | None:
        if key in header_map:
            return header_map[key]
        return default

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        title = row[col("название этапа")]
        if not title:
            errors.append(f"Строка {row_index}: пустое название этапа")
            continue

        order_cell = col("порядок")
        order_value = len(stages)
        if order_cell is not None and row[order_cell] is not None:
            try:
                order_value = int(row[order_cell])
            except ValueError:
                errors.append(f"Строка {row_index}: не удалось разобрать порядок '{row[order_cell]}'")
                continue

        status_cell = col("статус")
        status_raw = row[status_cell] if status_cell is not None else None
        status_value = StageStatus.NOT_STARTED
        if status_raw:
            normalized = str(status_raw).strip().lower()
            status_value = STATUS_ALIASES.get(normalized, None) or StageStatus.__members__.get(normalized.upper(), None)
            if status_value is None:
                errors.append(f"Строка {row_index}: неизвестный статус '{status_raw}'")
                continue

        risk_cell = col("риск")
        risk_value = _parse_bool(row[risk_cell]) if risk_cell is not None else False

        checklist_cell = col("чек-лист")
        checklist_raw = row[checklist_cell] if checklist_cell is not None else None
        checklist_items: list[str] = []
        if checklist_raw:
            checklist_items = [part.strip() for part in str(checklist_raw).split(";") if part.strip()]

        checklist_models = []
        for idx, item in enumerate(checklist_items):
            text = item
            done = False
            if item.startswith("[x]"):
                done = True
                text = item[3:].strip()
            elif item.startswith("[ ]"):
                text = item[3:].strip()
            checklist_models.append(ChecklistItem(title=text, done=done, order=idx))

        stage = GTMStage(
            title=str(title).strip(),
            description=(row[col("описание")] if col("описание") is not None else None),
            order=order_value,
            planned_start=row[col("плановая дата начала")] if col("плановая дата начала") is not None else None,
            planned_end=row[col("плановая дата окончания")] if col("плановая дата окончания") is not None else None,
            actual_end=row[col("фактическая дата завершения")] if col("фактическая дата завершения") is not None else None,
            status=status_value,
            risk_flag=risk_value,
            checklist=checklist_models,
        )

        stages.append(stage)

    stages.sort(key=lambda s: s.order)

    stage_index = {normalize(str(stage.title)): stage for stage in stages}
    stage_order_map = {stage.id: idx for idx, stage in enumerate(stages)}

    if "Задачи" in workbook.sheetnames:
        task_sheet = workbook["Задачи"]
        try:
            first_task_row = next(task_sheet.iter_rows(max_row=1))
        except StopIteration:
            first_task_row = []
        task_header = [str(cell.value).strip() if cell.value is not None else "" for cell in first_task_row]
        task_header_map = {title.lower(): idx for idx, title in enumerate(task_header) if title}

        def task_col(key: str, default: int | None = None) -> int | None:
            if key in task_header_map:
                return task_header_map[key]
            return default

        task_order_col = task_col("порядок задачи")
        for row_index, row in enumerate(task_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            stage_title = row[task_col("этап")]
            normalized_stage = normalize(stage_title if stage_title is not None else "")
            stage = stage_index.get(normalized_stage)
            if stage is None:
                errors.append(f"Строка {row_index} (Задачи): этап '{stage_title}' не найден среди импортируемых этапов")
                continue

            order_value = row[task_order_col] if task_order_col is not None else None
            task_order = 0
            try:
                task_order = int(order_value) if order_value not in (None, "") else 0
            except (TypeError, ValueError):
                errors.append(f"Строка {row_index} (Задачи): неверный порядок задачи '{order_value}'")
                continue

            title = row[task_col("название задачи")]
            if not title:
                errors.append(f"Строка {row_index} (Задачи): пустое название задачи")
                continue

            status_raw = row[task_col("статус")]
            status = TaskStatus.TODO
            if status_raw:
                status = TASK_STATUS_ALIASES.get(str(status_raw).strip().lower(), None) or TaskStatus.__members__.get(
                    str(status_raw).strip().upper(), TaskStatus.TODO
                )

            urgency_raw = row[task_col("срочность")]
            urgency = TaskUrgency.NORMAL
            if urgency_raw:
                urgency = URGENCY_ALIASES.get(str(urgency_raw).strip().lower(), TaskUrgency.NORMAL)

            important_raw = row[task_col("важная")]
            important = _parse_bool(important_raw) if important_raw is not None else False

            task_obj = Task(
                title=str(title).strip(),
                description=row[task_col("описание")],
                status=status,
                due_date=row[task_col("срок")],
                important=important,
                urgency=urgency,
                gtm_stage_id=stage.id,
                subtasks=[],
                comments=[],
            )
            tasks_raw.append((task_order, stage, task_obj))

    subtasks_map: dict[tuple[UUID, int], list[Subtask]] = {}
    if "Подзадачи" in workbook.sheetnames:
        sub_sheet = workbook["Подзадачи"]
        try:
            first_sub_row = next(sub_sheet.iter_rows(max_row=1))
        except StopIteration:
            first_sub_row = []
        sub_header = [str(cell.value).strip() if cell.value is not None else "" for cell in first_sub_row]
        sub_header_map = {title.lower(): idx for idx, title in enumerate(sub_header) if title}

        def sub_col(key: str, default: int | None = None) -> int | None:
            if key in sub_header_map:
                return sub_header_map[key]
            return default

        for row_index, row in enumerate(sub_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            stage_title = row[sub_col("этап")]
            normalized_stage = normalize(stage_title if stage_title is not None else "")
            stage = stage_index.get(normalized_stage)
            if stage is None:
                errors.append(f"Строка {row_index} (Подзадачи): этап '{stage_title}' не найден")
                continue

            task_order_raw = row[sub_col("порядок задачи")]
            try:
                task_order_value = int(task_order_raw) if task_order_raw not in (None, "") else 0
            except (TypeError, ValueError):
                errors.append(f"Строка {row_index} (Подзадачи): неверный порядок задачи '{task_order_raw}'")
                continue

            title = row[sub_col("название подзадачи")]
            if not title:
                errors.append(f"Строка {row_index} (Подзадачи): пустое название подзадачи")
                continue

            done_raw = row[sub_col("выполнена")]
            done = _parse_bool(done_raw)
            order_raw = row[sub_col("порядок подзадачи")]
            try:
                order_val = int(order_raw) if order_raw not in (None, "") else 0
            except (TypeError, ValueError):
                order_val = 0

            key = (stage.id, task_order_value)
            subtasks_map.setdefault(key, []).append(
                Subtask(title=str(title).strip(), done=done, order=order_val)
            )

    # Связываем подзадачи с задачами и сортируем по порядку этапов
    ordered_tasks: list[Task] = []
    for task_order, stage, task_obj in sorted(tasks_raw, key=lambda t: (stage_order_map.get(t[1].id, 999), t[0])):
        key = (stage.id, task_order)
        subs = sorted(subtasks_map.get(key, []), key=lambda s: s.order)
        task_obj.subtasks = subs
        ordered_tasks.append(task_obj)

    return stages, ordered_tasks, errors


PROJECT_STATUS_ALIASES = {
    "в работе": ProjectStatus.IN_PROGRESS,
    "in progress": ProjectStatus.IN_PROGRESS,
    "активный": ProjectStatus.IN_PROGRESS,
    "запущен": ProjectStatus.LAUNCHED,
    "launched": ProjectStatus.LAUNCHED,
    "закрытый": ProjectStatus.CLOSED,
    "closed": ProjectStatus.CLOSED,
    "eol": ProjectStatus.EOL,
    "архив": ProjectStatus.ARCHIVED,
    "архивный": ProjectStatus.ARCHIVED,
    "archived": ProjectStatus.ARCHIVED,
}

PRIORITY_ALIASES = {
    "низкий": PriorityLevel.LOW,
    "low": PriorityLevel.LOW,
    "средний": PriorityLevel.MEDIUM,
    "medium": PriorityLevel.MEDIUM,
    "высокий": PriorityLevel.HIGH,
    "high": PriorityLevel.HIGH,
}


def import_projects_from_excel(
    content: bytes,
    groups: Iterable[ProductGroup],
    existing_projects: Iterable[Project],
) -> tuple[list[Project], list[str]]:
    """Распарсить Excel со списком проектов и вернуть новые модели."""

    try:
        workbook = load_workbook(filename=BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        return [], [f"Не удалось прочитать Excel: {exc}"]

    sheet = workbook.active
    try:
        first_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        return [], ["Файл Excel пуст"]

    header_titles = [str(cell.value).strip() if cell.value is not None else "" for cell in first_row]
    header_map = {title.lower(): idx for idx, title in enumerate(header_titles) if title}
    original_titles = {title.lower(): title for title in header_titles if title}

    required_columns = {"название проекта", "продуктовая группа", "бренд", "статус"}
    missing = required_columns - set(header_map)
    if missing:
        return [], [f"Отсутствуют обязательные столбцы: {', '.join(sorted(missing))}"]

    existing_by_id = {str(p.id): p for p in existing_projects}
    group_by_name = {g.name.strip().lower(): g for g in groups}

    def col(key: str, default: int | None = None) -> int | None:
        if key in header_map:
            return header_map[key]
        return default

    def parse_status(raw) -> ProjectStatus:
        if raw is None:
            return ProjectStatus.IN_PROGRESS
        normalized = str(raw).strip().lower()
        return PROJECT_STATUS_ALIASES.get(normalized, ProjectStatus.__members__.get(normalized.upper(), ProjectStatus.IN_PROGRESS))

    def parse_priority(raw) -> PriorityLevel | None:
        if raw is None or raw == "":
            return None
        normalized = str(raw).strip().lower()
        return PRIORITY_ALIASES.get(normalized, PriorityLevel.__members__.get(normalized.upper(), None))

    def normalize_date(value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if value is None:
            return None
        text = str(value).strip()
        if "T" in text or " " in text:
            text = text.split("T")[0].split(" ")[0]
        for parser in (date.fromisoformat, lambda val: datetime.fromisoformat(val).date()):
            try:
                return parser(text)
            except ValueError:
                continue
        return None

    custom_fields_columns = [original_titles[key] for key in header_map if key.startswith(CUSTOM_FIELD_PREFIX.lower())]

    parsed: list[Project] = []
    errors: list[str] = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        name = row[col("название проекта")]
        if not name:
            errors.append(f"Строка {row_index}: не указано название проекта")
            continue

        group_name = row[col("продуктовая группа")]
        group = group_by_name.get(str(group_name).strip().lower()) if group_name else None
        if group is None:
            errors.append(f"Строка {row_index}: продуктовая группа '{group_name}' не найдена")
            continue

        status_raw = row[col("статус")]
        status = parse_status(status_raw)

        project_id_val = row[col("id")] if col("id") is not None else None
        project_id = str(project_id_val).strip() if project_id_val else None
        existing = existing_by_id.get(project_id) if project_id else None

        base_kwargs = dict(
            group_id=group.id,
            name=str(name),
            brand=str(row[col("бренд")]) if row[col("бренд")] is not None else "",
            market=row[col("рынок/регион")],
            status=status,
            planned_launch=normalize_date(row[col("плановая дата запуска")]),
            actual_launch=normalize_date(row[col("фактическая дата запуска")]),
            current_gtm_stage_id=None,
            priority=parse_priority(row[col("приоритет")]),
            moq=_coerce_value(row[col("moq")], FieldType.NUMBER) if col("moq") is not None else None,
            fob_price=_coerce_value(row[col("fob")], FieldType.NUMBER) if col("fob") is not None else None,
            promo_price=_coerce_value(row[col("promo")], FieldType.NUMBER) if col("promo") is not None else None,
            rrp_price=_coerce_value(row[col("rrp")], FieldType.NUMBER) if col("rrp") is not None else None,
            short_description=row[col("краткое описание")],
            full_description=row[col("полное описание")],
            custom_fields={},
        )

        current_stage_title = row[col("текущий gtm-этап")] if col("текущий gtm-этап") is not None else None
        if existing and current_stage_title:
            matched_stage = next(
                (stage for stage in existing.gtm_stages if stage.title.strip().lower() == str(current_stage_title).strip().lower()),
                None,
            )
            base_kwargs["current_gtm_stage_id"] = matched_stage.id if matched_stage else None

        for cf_col in custom_fields_columns:
            raw_value = row[header_map[cf_col.lower()]]
            if raw_value in (None, ""):
                continue
            key = cf_col[len(CUSTOM_FIELD_PREFIX) :]
            base_kwargs["custom_fields"][key] = raw_value

        short_id_val = row[col("короткий id")]
        if existing:
            updated = existing.model_copy(update=base_kwargs)
            if short_id_val:
                try:
                    updated.short_id = int(short_id_val)
                except (TypeError, ValueError):
                    pass
            parsed.append(updated)
        else:
            parsed.append(
                Project(
                    id=UUID(project_id) if project_id else uuid4(),
                    short_id=int(short_id_val) if isinstance(short_id_val, (int, float)) else None,
                    gtm_stages=[],
                    tasks=[],
                    characteristics=[],
                    files=[],
                    images=[],
                    comments=[],
                    history=[],
                    **base_kwargs,
                )
            )

    return parsed, errors


def _parse_project_sheet(
    sheet, groups: Iterable[ProductGroup], existing_project: Project
) -> tuple[Project, list[str], str | None]:
    try:
        first_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        return existing_project, ["Файл Excel пуст"], None

    header_titles = [str(cell.value).strip() if cell.value is not None else "" for cell in first_row]
    header_map = {title.lower(): idx for idx, title in enumerate(header_titles) if title}
    original_titles = {title.lower(): title for title in header_titles if title}

    required_columns = {"название проекта", "продуктовая группа"}
    missing = required_columns - set(header_map)
    if missing:
        return existing_project, [f"Отсутствуют обязательные столбцы: {', '.join(sorted(missing))}"], None

    custom_fields_columns = [original_titles[key] for key in header_map if key.startswith(CUSTOM_FIELD_PREFIX.lower())]
    group_by_name = {g.name.strip().lower(): g for g in groups}
    errors: list[str] = []

    def col(key: str, default: int | None = None) -> int | None:
        if key in header_map:
            return header_map[key]
        return default

    def parse_status(raw) -> ProjectStatus:
        if raw is None:
            return existing_project.status
        normalized = str(raw).strip().lower()
        return PROJECT_STATUS_ALIASES.get(normalized, ProjectStatus.__members__.get(normalized.upper(), existing_project.status))

    def parse_priority(raw) -> PriorityLevel | None:
        if raw is None or raw == "":
            return existing_project.priority
        normalized = str(raw).strip().lower()
        return PRIORITY_ALIASES.get(normalized, PriorityLevel.__members__.get(normalized.upper(), existing_project.priority))

    def normalize_date(value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if value is None:
            return None
        text = str(value).strip()
        if "T" in text or " " in text:
            text = text.split("T")[0].split(" ")[0]
        for parser in (date.fromisoformat, lambda val: datetime.fromisoformat(val).date()):
            try:
                return parser(text)
            except ValueError:
                continue
        return None

    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    if not data_rows:
        return existing_project, ["Не найдены данные проекта"], None
    row = data_rows[0]

    name = row[col("название проекта")]
    if not name:
        errors.append("Не указано название проекта")

    group_cell = row[col("продуктовая группа")]
    group = group_by_name.get(str(group_cell).strip().lower()) if group_cell else None
    if group is None:
        errors.append(f"Продуктовая группа '{group_cell}' не найдена")

    current_stage_title = row[col("текущий gtm-этап")] if col("текущий gtm-этап") is not None else None

    base_kwargs = dict(
        name=str(name).strip() if name else existing_project.name,
        group_id=group.id if group else existing_project.group_id,
        brand=str(row[col("бренд")]) if col("бренд") is not None and row[col("бренд")] is not None else existing_project.brand,
        market=row[col("рынок/регион")]
        if col("рынок/регион") is not None
        else existing_project.market,
        status=parse_status(row[col("статус")]) if col("статус") is not None else existing_project.status,
        planned_launch=normalize_date(row[col("плановая дата запуска")])
        if col("плановая дата запуска") is not None
        else existing_project.planned_launch,
        actual_launch=normalize_date(row[col("фактическая дата запуска")])
        if col("фактическая дата запуска") is not None
        else existing_project.actual_launch,
        current_gtm_stage_id=existing_project.current_gtm_stage_id,
        priority=parse_priority(row[col("приоритет")]) if col("приоритет") is not None else existing_project.priority,
        moq=_coerce_value(row[col("moq")], FieldType.NUMBER)
        if col("moq") is not None
        else existing_project.moq,
        fob_price=_coerce_value(row[col("fob")], FieldType.NUMBER)
        if col("fob") is not None
        else existing_project.fob_price,
        promo_price=_coerce_value(row[col("promo")], FieldType.NUMBER)
        if col("promo") is not None
        else existing_project.promo_price,
        rrp_price=_coerce_value(row[col("rrp")], FieldType.NUMBER)
        if col("rrp") is not None
        else existing_project.rrp_price,
        short_description=row[col("краткое описание")]
        if col("краткое описание") is not None
        else existing_project.short_description,
        full_description=row[col("полное описание")]
        if col("полное описание") is not None
        else existing_project.full_description,
        custom_fields=dict(existing_project.custom_fields),
    )

    for cf_col in custom_fields_columns:
        raw_value = row[header_map[cf_col.lower()]]
        if raw_value in (None, ""):
            continue
        key = cf_col[len(CUSTOM_FIELD_PREFIX) :]
        base_kwargs["custom_fields"][key] = raw_value

    short_id_val = row[col("короткий id")]
    updated = existing_project.model_copy(update=base_kwargs)
    if short_id_val not in (None, ""):
        try:
            updated.short_id = int(short_id_val)
        except (TypeError, ValueError):
            errors.append("Некорректный Короткий ID")

    return updated, errors, str(current_stage_title).strip() if current_stage_title else None


def import_project_bundle_from_excel(
    content: bytes, groups: Iterable[ProductGroup], existing_project: Project
) -> tuple[Project, list[str]]:
    try:
        workbook = load_workbook(filename=BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        return existing_project, [f"Не удалось прочитать Excel: {exc}"]

    errors: list[str] = []

    basics_sheet = workbook["Основные параметры"] if "Основные параметры" in workbook.sheetnames else workbook.active
    project_core, core_errors, current_stage_title = _parse_project_sheet(basics_sheet, groups, existing_project)
    errors.extend(core_errors)

    gtm_sheet = workbook["GTM"] if "GTM" in workbook.sheetnames else None
    stages, tasks, gtm_errors = import_gtm_stages_from_sheet(gtm_sheet) if gtm_sheet else ([], [], ["Лист GTM не найден"])
    errors.extend(gtm_errors)

    characteristics_sheet = (
        workbook["Характеристики"] if "Характеристики" in workbook.sheetnames else None
    )
    if characteristics_sheet:
        char_sections, char_errors, _ = import_characteristics_from_sheet(characteristics_sheet, project_core)
        errors.extend(char_errors)
    else:
        char_sections, _ = project_core.characteristics, None
        errors.append("Лист характеристик не найден")

    if errors:
        return project_core, errors

    if current_stage_title:
        matched_stage = next(
            (s for s in stages if s.title and s.title.strip().lower() == current_stage_title.strip().lower()),
            None,
        )
        project_core.current_gtm_stage_id = matched_stage.id if matched_stage else None

    updated = project_core.model_copy(update={"gtm_stages": stages, "tasks": tasks, "characteristics": char_sections})
    return updated, errors


def export_all_characteristics(
    projects: Iterable[Project],
    groups: Iterable[ProductGroup],
    project_filter: set[UUID] | None = None,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    used_names: set[str] = set()
    group_lookup = {g.id: g.name for g in groups}
    exported_any = False

    for project in projects:
        if project_filter and project.id not in project_filter:
            continue
        sheet_name = _make_sheet_name(project.name or "Проект", used_names)
        _write_characteristics_sheet(workbook, project, sheet_name)
        sheet = workbook[sheet_name]
        sheet.insert_rows(1)
        sheet["A1"] = f"Проект: {project.name}"
        sheet["B1"] = f"Группа: {group_lookup.get(project.group_id, '')}"
        exported_any = True

    if not exported_any:
        sheet = workbook.create_sheet("Характеристики")
        sheet.append(["Нет проектов для экспорта"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def import_characteristics_bulk(
    content: bytes, projects: Iterable[Project]
) -> tuple[dict[UUID, list[CharacteristicSection]], list[str]]:
    try:
        workbook = load_workbook(filename=BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        return {}, [f"Не удалось прочитать Excel: {exc}"]

    project_index = {p.name.strip().lower(): p for p in projects}
    updates: dict[UUID, list[CharacteristicSection]] = {}
    errors: list[str] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        project = project_index.get(sheet_name.strip().lower())
        if project is None:
            errors.append(f"Лист '{sheet_name}': проект не найден")
            continue
        sections, section_errors, _ = import_characteristics_from_sheet(sheet, project)
        errors.extend([f"{sheet_name}: {err}" for err in section_errors])
        if not section_errors:
            updates[project.id] = sections

    return updates, errors


FIELD_TYPE_ALIASES = {
    "text": FieldType.TEXT,
    "текст": FieldType.TEXT,
    "number": FieldType.NUMBER,
    "число": FieldType.NUMBER,
    "select": FieldType.SELECT,
    "list": FieldType.SELECT,
    "список": FieldType.SELECT,
    "checkbox": FieldType.CHECKBOX,
    "флажок": FieldType.CHECKBOX,
    "галочка": FieldType.CHECKBOX,
    "other": FieldType.OTHER,
    "другое": FieldType.OTHER,
}


def _coerce_value(raw: str | int | float | bool | None, field_type: FieldType) -> str | int | float | bool | None:
    if raw is None:
        return None

    if field_type == FieldType.CHECKBOX:
        return _parse_bool(raw)

    if field_type == FieldType.NUMBER:
        try:
            # Excel числа уже приходят числовыми; строки аккуратно конвертируем
            if isinstance(raw, (int, float)):
                return raw
            cleaned = str(raw).strip().replace(",", ".")
            numeric = float(cleaned)
            return int(numeric) if numeric.is_integer() else numeric
        except (ValueError, TypeError):
            return raw

    return raw


def import_characteristics_from_excel(
    content: bytes, project: Project
) -> tuple[list[CharacteristicSection], list[str], dict[str, int]]:
    """Распарсить Excel с характеристиками и вернуть обновлённые секции, ошибки и отчёт."""

    try:
        workbook = load_workbook(filename=BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        return [], [f"Не удалось прочитать Excel: {exc}"]

    sheet = workbook.active
    return import_characteristics_from_sheet(sheet, project)


def import_characteristics_from_sheet(
    sheet, project: Project
) -> tuple[list[CharacteristicSection], list[str], dict[str, int]]:
    try:
        first_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        return [], ["Файл Excel пуст"]

    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in first_row]
    header_map = {title.lower(): idx for idx, title in enumerate(header_row) if title}

    required_columns = {"секция", "label ru", "label en", "value ru", "value en", "тип поля"}
    missing = required_columns - set(header_map)
    if missing:
        return [], [f"Отсутствуют обязательные столбцы: {', '.join(sorted(missing))}"]

    errors: list[str] = []
    sections_copy: list[CharacteristicSection] = [section.model_copy(deep=True) for section in project.characteristics]
    report = {
        "sections_created": 0,
        "fields_created": 0,
        "fields_updated": 0,
        "rows_skipped": 0,
    }

    # Быстрый доступ к существующим секциям и порядкам, чтобы корректно добавлять новые
    section_index = {normalize(section.title): section for section in sections_copy}
    max_section_order = max((section.order for section in sections_copy), default=-1)

    def col(key: str, default: int | None = None) -> int | None:
        if key in header_map:
            return header_map[key]
        return default

    def normalize(value: str | None) -> str:
        return value.strip().lower() if value else ""

    def parse_field_type(raw: str | None, fallback: FieldType) -> FieldType:
        if raw is None:
            return fallback
        normalized = normalize(str(raw))
        return FIELD_TYPE_ALIASES.get(normalized, fallback)

    section_order_col = col("порядок секции")
    field_order_col = col("порядок поля")

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        section_title = row[col("секция")]
        if not section_title:
            errors.append(f"Строка {row_index}: не заполнено название секции")
            report["rows_skipped"] += 1
            continue

        label_ru = row[col("label ru")]
        label_en = row[col("label en")]
        if not label_ru and not label_en:
            errors.append(f"Строка {row_index}: не указаны Label RU/EN")
            report["rows_skipped"] += 1
            continue

        normalized_section = normalize(str(section_title))
        section = section_index.get(normalized_section)
        if section is None:
            max_section_order += 1
            order_value = row[section_order_col] if section_order_col is not None else None
            try:
                order = int(order_value) if order_value not in (None, "") else max_section_order
            except (TypeError, ValueError):
                order = max_section_order
            section = CharacteristicSection(title=str(section_title), order=order, fields=[])
            sections_copy.append(section)
            section_index[normalized_section] = section
            report["sections_created"] += 1

        max_field_order = max((fld.order for fld in section.fields), default=-1)
        normalized_ru = normalize(label_ru or "")
        normalized_en = normalize(label_en or "")
        field = next(
            (
                f
                for f in section.fields
                if normalize(f.label_ru) == normalized_ru and normalize(f.label_en) == normalized_en
            ),
            None,
        )

        type_cell_idx = col("тип поля")
        provided_type = parse_field_type(row[type_cell_idx], FieldType.TEXT) if type_cell_idx is not None else FieldType.TEXT

        if field is None:
            order_value = row[field_order_col] if field_order_col is not None else None
            try:
                order = int(order_value) if order_value not in (None, "") else max_field_order + 1
            except (TypeError, ValueError):
                order = max_field_order + 1

            field = CharacteristicField(
                label_ru=str(label_ru or label_en or ""),
                label_en=str(label_en or label_ru or ""),
                field_type=provided_type,
                order=order,
            )
            section.fields.append(field)
            report["fields_created"] += 1
        else:
            field.field_type = parse_field_type(row[type_cell_idx], field.field_type) if type_cell_idx is not None else field.field_type
            report["fields_updated"] += 1

        value_ru_idx = col("value ru")
        value_en_idx = col("value en")
        if value_ru_idx is not None:
            field.value_ru = _coerce_value(row[value_ru_idx], field.field_type)
        if value_en_idx is not None:
            field.value_en = _coerce_value(row[value_en_idx], field.field_type)

    sections_copy.sort(key=lambda s: s.order)
    for section in sections_copy:
        section.fields.sort(key=lambda f: f.order)

    return sections_copy, errors, report
